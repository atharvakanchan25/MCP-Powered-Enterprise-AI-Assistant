from pathlib import Path
from openpyxl import load_workbook
from app.rag.parsers.base import ParsedChunk


class XlsxParser:
    def parse(self, path: Path) -> list[ParsedChunk]:
        wb = load_workbook(str(path), read_only=True, data_only=True)
        chunks = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                row_text = "\t".join("" if v is None else str(v) for v in row)
                if row_text.strip():
                    rows.append(row_text)
            if rows:
                chunks.append(ParsedChunk("\n".join(rows), {"sheet": sheet_name}))
        wb.close()
        return chunks
